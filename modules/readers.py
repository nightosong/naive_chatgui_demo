import logging
import openpyxl
import pandas as pd
from typing import Any, Dict
from paddleocr import PaddleOCR

logging.getLogger("ppocr").setLevel(level=logging.WARNING)


def read_image_text(_file: Any, _more=False) -> list:
    """读取图片文字
    _file: Image for OCR. An ndarray, img_path, or a list of ndarrays.
    _more: Show more information
    """
    ocr = PaddleOCR(use_angle_cls=True, lang="ch")
    results = ocr.ocr(_file, cls=True)
    if not _more:
        return [line[1][0] for line in results]
    return results


def read_excel(_file: Any) -> list:
    df = pd.read_excel(_file)
    return df.values.tolist()


class ExcelReader:
    @staticmethod
    def insert_cell(xlsx_file, row, column, value, sheet_name=None):
        workbook = openpyxl.load_workbook(xlsx_file)
        if not sheet_name:
            sheet = workbook.active
        else:
            sheet = workbook[sheet_name]
        sheet.cell(row, column, value)
        workbook.save(xlsx_file)

    @staticmethod
    def load_sheet_data(xlsx_file) -> Dict[str, list]:
        workbook = openpyxl.load_workbook(xlsx_file)
        sheet_data = {}
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            headers = [cell.value for cell in sheet[1]]
            values = list(sheet.iter_rows(min_row=2, values_only=True))
            sheet_data[sheet_name] = {"headers": headers, "values": values}
        return sheet_data

    @staticmethod
    def load_data_by_row(xlsx_file):
        workbook = openpyxl.load_workbook(xlsx_file)
        sheet = workbook.active
        yield from sheet.iter_rows(min_row=2, values_only=True)
        workbook.close()


def clear_excel():
    empty_df = pd.DataFrame()

    # 写入 Excel 文件（会清空文件内容）
    empty_df.to_excel("result.xlsx", index=False)

    print("Excel file has been cleared.")
